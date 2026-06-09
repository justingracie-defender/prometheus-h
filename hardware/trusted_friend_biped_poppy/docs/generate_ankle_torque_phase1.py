from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = Path('/home/ubuntu/prometheus-h/hardware/trusted_friend_biped_poppy/docs')
CSV_PATH = OUT_DIR / 'ankle_torque_phase1.csv'
XLSX_PATH = OUT_DIR / 'ankle_torque_phase1.xlsx'

G = 9.81
DYNAMIC_MULTIPLIER = 1.25
SAFETY_FACTOR = 2.0
GATE_NM = 110.0

rows = []
for mass_kg in [18, 20, 22]:
    for com_m in [0.15, 0.20, 0.25]:
        static_nm = mass_kg * G * com_m
        required_nm = static_nm * DYNAMIC_MULTIPLIER * SAFETY_FACTOR
        rows.append({
            'case': f'{mass_kg} kg @ {com_m:.2f} m COM offset',
            'robot_mass_kg': mass_kg,
            'ankle_to_com_offset_m': com_m,
            'gravity_m_s2': G,
            'dynamic_multiplier_shuffle': DYNAMIC_MULTIPLIER,
            'safety_factor': SAFETY_FACTOR,
            'static_torque_nm': round(static_nm, 1),
            'required_design_torque_nm': round(required_nm, 1),
            'phase1_gate_nm': GATE_NM,
            'gate_result': 'PASS threshold reached' if required_nm >= GATE_NM else 'Below 110 Nm threshold but still must be validated',
        })

fieldnames = list(rows[0].keys())
with CSV_PATH.open('w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

wb = Workbook()
ws = wb.active
ws.title = 'Ankle Phase 1 Torque'
ws.append(fieldnames)
for row in rows:
    ws.append([row[k] for k in fieldnames])

header_fill = PatternFill('solid', fgColor='1F4E79')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for col in range(1, len(fieldnames) + 1):
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = max(14, min(34, max(len(str(ws.cell(row=r, column=col).value)) for r in range(1, ws.max_row + 1)) + 2))

ws.freeze_panes = 'A2'

summary = wb.create_sheet('Actuator Shortlist')
summary_rows = [
    ['Candidate', 'Rated / Continuous Torque (Nm)', 'Peak Torque (Nm)', 'Mass (g)', 'Phase 1 Role', 'Preliminary Decision'],
    ['CubeMars AKH70-48 V1.0 KV41', 74, 222, 1396, 'Primary ankle CAD envelope and fixture candidate', 'Shortlist first; peak exceeds 110 Nm, rated does not'],
    ['CubeMars AK80-64 KV80', 48, 120, 850, 'Lightweight fallback candidate', 'Shortlist second; barely clears 110 Nm only on peak'],
    ['ROBOTIS H54-200-S500-R + external 3:1 reduction', 134.1, None, '855 + gearbox', 'ROBOTIS-family alternative', 'Possible if gearbox, backlash, speed, and bracket loads pass'],
    ['ROBOTIS MX-64 / XM430 class', 6.0, None, '82-135', 'Not primary ankle actuator', 'Reject for main ankle support without major external transmission'],
]
for r in summary_rows:
    summary.append(r)
for cell in summary[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
for col in range(1, 7):
    letter = get_column_letter(col)
    summary.column_dimensions[letter].width = 28

wb.save(XLSX_PATH)
print(CSV_PATH)
print(XLSX_PATH)
